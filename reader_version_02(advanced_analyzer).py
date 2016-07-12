#Author: Joseph Mammo | Joseph.mammo@coyotes.usd.edu
#Project: Radon, Temperature, Air Pressure, and Humidity data analyzer; Analyzes daily, monthly and yearly average, min, max and average error for the 4850 Davis Campus radon detector data.

import openpyxl
from datetime import datetime
import numpy as np
import csv

print("** Initializing ...")
wb = openpyxl.load_workbook('SURF-RadonTrends.xlsx') #Todo: change the name to a custom Excel file
print("** Loaded xlsx file to the memory ...")
sheet = wb.get_sheet_by_name('1250L-Ross') # TODO: change the name to the specific sheet
print("** Opened 4850L Ross Campus sheet ...")
print("** Max sheets: " + str(sheet.max_row))
print("** Max columns: " + str(sheet.max_column))
	
Date = [] #A
Rn_hourly = [] #D
Temperature = [] #K
Air_Pressure = [] #L
Humidity = [] #M

print("** Gathering and storing data ...")

#Get the date data and append it to Data[]
for rowOfCellObjects in sheet['A1' : 'A' + str(sheet.max_row)]:
    for cellObj in rowOfCellObjects:
        if (cellObj.value != None):
            #d = cellObj.value.isocalendar()[2]
            if(isinstance(cellObj.value, datetime)):
                Date.append(cellObj.value)
            else:
                Date.append(datetime.strptime(cellObj.value, '%m/%d/%Y'))
            #Date.append(cellObj.value.strftime("%Y-%m-%d %H:%M"))
        #print(cellObj.coordinate, cellObj.value)

#Get the Radon data and append it to Rn_hourly[] 
for rowOfCellObjects in sheet['D1' : 'D' + str(sheet.max_row)]:
    for cellObj in rowOfCellObjects:
        if(isinstance(cellObj.value, long)):
            Rn_hourly.append(cellObj.value)
        #print(cellObj.coordinate, cellObj.value)

#Get the temperature data and append it Temperature[]
for rowOfCellObjects in sheet['K1' : 'K' + str(sheet.max_row)]:
    for cellObj in rowOfCellObjects:
        if(cellObj.value != None):
        #if isinstance(cellObj.value, float):
            Temperature.append(cellObj.value)
        #print(cellObj.coordinate, cellObj.value)

#Get the air pressure data and append it to Air_pressure[]
for rowOfCellObjects in sheet['L1' : 'L' + str(sheet.max_row)]:
    for cellObj in rowOfCellObjects:
        if(cellObj.value != None):
            if isinstance(cellObj.value, long):
                Air_Pressure.append(cellObj.value)
        #print(cellObj.coordinate, cellObj.value)
        
#Get the humidity data and append it to Humidity[]
for rowOfCellObjects in sheet['M1' : 'M' + str(sheet.max_row)]:
    for cellObj in rowOfCellObjects:
        if(cellObj.value != None):
            if isinstance(cellObj.value, float) or isinstance(cellObj.value, long):
                Humidity.append(cellObj.value)
        #print(cellObj.coordinate, cellObj.value)
        
print("** Checking input values with thier corresponding day input ...  ")
print("** Day count: " + str(len(Date)))
print("** Temperature count: " + str(len(Temperature)))
print("** Radon Count: " + str(len(Rn_hourly)))
print("** Air pressure count: " + str(len(Air_Pressure)))
print("** Humidity count: " + str(len(Humidity)))

if((str(len(Rn_hourly)) != str(len(Date))) or (str(len(Temperature)) != str(len(Date))) or (str(len(Air_Pressure)) != str(len(Date))) or (str(len(Humidity)) != str(len(Date)))):
    print("---------------------------------------------------------------------------")
    print("** Just a heads up, an error might occur")
    print("---------------------------------------------------------------------------")
    
avg_final = []
dummy_avg = []
dummy_day = []
dummy_min =[]
dummy_max = []
dummy_err = []
dumyy_dummy = []

def analyzer_daily(obj):
    count = -1 
    ind = Date[0].day
    dummy_day.append(Date[0].strftime("%Y-%m-%d"))
    for row in Date:
        count = count + 1
        dayye = row.day  
        if(dayye == ind):
            if row.strftime("%Y-%m-%d") not in dummy_day:
                dummy_day.append(row.strftime("%Y-%m-%d"))
            count_two = count - 2
            #print(str(row) + " : " + str(obj[count_two]))
            count_two = count_two + 1
            dummy_avg.append(obj[count_two])
        else:
            dummy_avg.append(obj[count_two+1])
            avg_final.append(np.mean(dummy_avg)) # calculate the average
            dummy_max.append(np.max(dummy_avg)) # get the max value
            dummy_min.append(np.min(dummy_avg)) # get the min value
            error = (np.std(dummy_avg))/(np.sqrt(len(dummy_avg) - 1)) # get the error
            dummy_err.append(error)
            dummy_day.append(row.strftime("%Y-%m-%d"))
            print("avg: " + str(np.mean(dummy_avg)))
            print("min: " + str(np.min(dummy_avg)))
            print("max: " + str(np.max(dummy_avg)))
            print("error: " + str(error))            
            del dummy_avg[:] #flush out existing data
            ind = dayye
            print("---------------------------- End day data -----------------------")
    return

def analyzer_monthly(obj):
    count = -1 
    ind = Date[0].month
    dummy_day.append(Date[0].strftime("%Y-%m"))
    for row in Date:
        count = count + 1
        dayye = row.month  
        if(dayye == ind):
            if row.strftime("%Y-%m") not in dummy_day:
                dummy_day.append(row.strftime("%Y-%m"))
            count_two = count - 2
            #print(str(row) + " : " + str(obj[count_two]))
            count_two = count_two + 1
            dummy_avg.append(obj[count_two])
        else:
            dummy_avg.append(obj[count_two+1])
            avg_final.append(np.mean(dummy_avg)) # calculate the average
            dummy_max.append(np.max(dummy_avg)) # get the max value
            dummy_min.append(np.min(dummy_avg)) # get the min value
            error = (np.std(dummy_avg))/(np.sqrt(len(dummy_avg) - 1)) # get the error
            dummy_err.append(error)
            dummy_day.append(row.strftime("%Y-%m"))
            print("avg: " + str(np.mean(dummy_avg)))
            print("min: " + str(np.min(dummy_avg)))
            print("max: " + str(np.max(dummy_avg)))
            print("error: " + str(error))             
            del dummy_avg[:] #flush out existing data
            ind = dayye
            print("---------------------------- End month data -----------------------")
    return

def analyzer_weekly(obj):
    count = -1 
    ind = Date[0].isocalendar()[1]
    dummy_day.append(Date[0].strftime("%W"))
    dumyy_dummy.append(Date[0].strftime("%Y-%m-%d"))
    for row in Date:
        count = count + 1
        dayye = row.isocalendar()[1]  
        if(dayye == ind):
            if row.strftime("%W") not in dummy_day:
                dummy_day.append(row.strftime("%W"))
            #if row.strftime("%Y-%m-%d") not in dumyy_dummy:
                #dumyy_dummy.append(row.strftime("%Y-%m-%d"))
            count_two = count - 2
            #print(str(row) + " : " + str(obj[count_two]))
            count_two = count_two + 1
            dummy_avg.append(obj[count_two])
        else:
            dummy_avg.append(obj[count_two+1])
            avg_final.append(np.mean(dummy_avg)) # calculate the average
            dummy_max.append(np.max(dummy_avg)) # get the max value
            dummy_min.append(np.min(dummy_avg)) # get the min value
            error = (np.std(dummy_avg))/(np.sqrt(len(dummy_avg) - 1)) # get the error
            dummy_err.append(error)
            dummy_day.append(row.strftime("%W"))
            dumyy_dummy.append(row.strftime("%Y-%m-%d"))
            print("week end date: " + str(row.strftime("%Y-%m-%d")))
            print("avg: " + str(np.mean(dummy_avg)))
            print("min: " + str(np.min(dummy_avg)))
            print("max: " + str(np.max(dummy_avg)))
            print("error: " + str(error))             
            del dummy_avg[:] #flush out existing data
            ind = dayye
            print("---------------------------- End week data -----------------------")
    return

def flushout():
    print("** Flshing out older data ...")
    del avg_final[:]
    del dummy_day[:]
    del dummy_min[:]
    del dummy_max[:]
    del dummy_err[:]  
    del dumyy_dummy[:]
    return
# call the analyzer functions

print("** ANALYZING DAILY VALUES ...")
print("** Analyzing Radon Data ...")
analyzer_daily(Rn_hourly)

print("** Writing analyzed data to 'Rn.csv' ...")
with open('Rn_daily.csv', 'w') as f:
    writer = csv.writer(f)
    writer.writerows(zip(dummy_day,avg_final,dummy_min,dummy_max,dummy_err))
    
flushout()

print("** Analyzing Temperature Data ...")
analyzer_daily(Temperature)

print("** Writing analyzed data to 'Temperature.csv' ...")
with open('Temperature_daily.csv', 'w') as f:
    writer = csv.writer(f)
    writer.writerows(zip(dummy_day,avg_final,dummy_min,dummy_max,dummy_err))
    
flushout()

print("** Analyzing Air Pressure Data ...")
analyzer_daily(Air_Pressure)

print("** Writing analyzed data to 'Air_pressure.csv' ...")
with open('Air_pressure_daily.csv', 'w') as f:
    writer = csv.writer(f)
    writer.writerows(zip(dummy_day,avg_final,dummy_min,dummy_max,dummy_err))
    
flushout()

print("** Analyzing Humidity Data ...")
analyzer_daily(Humidity)

print("** Writing analyzed data to 'Humidity.csv' ...")
with open('Humidity_daily.csv', 'w') as f:
    writer = csv.writer(f)
    writer.writerows(zip(dummy_day,avg_final,dummy_min,dummy_max,dummy_err))
    
flushout()

print("** ANALYZING WEEKLY VALUES ...")
print("** Analyzing Radon Data ...")
analyzer_weekly(Rn_hourly)

print("** Writing analyzed data to 'Rn.csv' ...")
with open('Rn_weekly.csv', 'w') as f:
    writer = csv.writer(f)
    writer.writerows(zip(dummy_day,avg_final,dummy_min,dummy_max,dummy_err,dumyy_dummy))
    
flushout()

print("** Analyzing Temperature Data ...")
analyzer_weekly(Temperature)

print("** Writing analyzed data to 'Temperature.csv' ...")
with open('Temperature_weekly.csv', 'w') as f:
    writer = csv.writer(f)
    writer.writerows(zip(dummy_day,avg_final,dummy_min,dummy_max,dummy_err,dumyy_dummy))
    
flushout()

print("** Analyzing Air Pressure Data ...")
analyzer_weekly(Air_Pressure)

print("** Writing analyzed data to 'Air_pressure.csv' ...")
with open('Air_pressure_weekly.csv', 'w') as f:
    writer = csv.writer(f)
    writer.writerows(zip(dummy_day,avg_final,dummy_min,dummy_max,dummy_err,dumyy_dummy))
    
flushout()

print("** Analyzing Humidity Data ...")
analyzer_weekly(Humidity)

print("** Writing analyzed data to 'Humidity.csv' ...")
with open('Humidity_weekly.csv', 'w') as f:
    writer = csv.writer(f)
    writer.writerows(zip(dummy_day,avg_final,dummy_min,dummy_max,dummy_err,dumyy_dummy))
    
flushout()

print("** ANALYZING MONTHLY VALUES ...")
print("** Analyzing Radon Data ...")
analyzer_monthly(Rn_hourly)

print("** Writing analyzed data to 'Rn.csv' ...")
with open('Rn_monthly.csv', 'w') as f:
    writer = csv.writer(f)
    writer.writerows(zip(dummy_day,avg_final,dummy_min,dummy_max,dummy_err))
    
flushout()

print("** Analyzing Temperature Data ...")
analyzer_monthly(Temperature)

print("** Writing analyzed data to 'Temperature.csv' ...")
with open('Temperature_monthly.csv', 'w') as f:
    writer = csv.writer(f)
    writer.writerows(zip(dummy_day,avg_final,dummy_min,dummy_max,dummy_err))
    
flushout()

print("** Analyzing Air Pressure Data ...")
analyzer_monthly(Air_Pressure)

print("** Writing analyzed data to 'Air_pressure.csv' ...")
with open('Air_pressure_monthly.csv', 'w') as f:
    writer = csv.writer(f)
    writer.writerows(zip(dummy_day,avg_final,dummy_min,dummy_max,dummy_err))
    
flushout()

print("** Analyzing Humidity Data ...")
analyzer_monthly(Humidity)

print("** Writing analyzed data to 'Humidity.csv' ...")
with open('Humidity_monthly.csv', 'w') as f:
    writer = csv.writer(f)
    writer.writerows(zip(dummy_day,avg_final,dummy_min,dummy_max,dummy_err))
    
flushout()

print("** Data analysis completed!")
