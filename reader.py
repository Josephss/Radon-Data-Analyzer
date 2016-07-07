#Author: Joseph Mammo | Joseph.mammo@coyotes.usd.edu
#Project: Radon, Temperature, Air Pressure, and Humidity data analyzer; Analyzes daily, monthly and yearly average, min, max and average error for the aforementioned data.

import openpyxl
from datetime import datetime
import numpy as np
import csv

print("** Initializing ...")
wb = openpyxl.load_workbook('SURF-RadonTrends.xlsx')
print("** Loaded xlsx file to the memory ...")
sheet = wb.get_sheet_by_name('4850L Davis Campus')
print("** Opened 4850L Davis Campus sheet ...")
print("** Max sheets: " + str(sheet.max_row))
print("** Max columns: " + str(sheet.max_column))
	
Date = [] #A
Rn_hourly = [] #D
Temperature = [] #K
Air_Pressure = [] #L
Humidity = [] #M

print("** Gathering and storing data ...")

#Get the date data and append it to Data[]
for rowOfCellObjects in sheet['A1' : 'A' + str(sheet.max_row)]:
    for cellObj in rowOfCellObjects:
        if (cellObj.value != None):
            #d = cellObj.value.isocalendar()[2]
            if(isinstance(cellObj.value, datetime)):
                Date.append(cellObj.value)
            else:
                Date.append(datetime.strptime(cellObj.value, '%m/%d/%Y'))
            #Date.append(cellObj.value.strftime("%Y-%m-%d %H:%M"))
        #print(cellObj.coordinate, cellObj.value)

#Get the Radon data and append it to Rn_hourly[] 
for rowOfCellObjects in sheet['D1' : 'D' + str(sheet.max_row)]:
    for cellObj in rowOfCellObjects:
        if(isinstance(cellObj.value, long)):
            Rn_hourly.append(cellObj.value)
        #print(cellObj.coordinate, cellObj.value)

#Get the temperature data and append it Temperature[]
for rowOfCellObjects in sheet['K1' : 'K' + str(sheet.max_row)]:
    for cellObj in rowOfCellObjects:
        if(cellObj.value != None):
        #if isinstance(cellObj.value, float):
            Temperature.append(cellObj.value)
        #print(cellObj.coordinate, cellObj.value)

#Get the air pressure data and append it to Air_pressure[]
for rowOfCellObjects in sheet['L1' : 'L' + str(sheet.max_row)]:
    for cellObj in rowOfCellObjects:
        if(cellObj.value != None):
        #if isinstance(cellObj.value, long):
            Air_Pressure.append(cellObj.value)
        #print(cellObj.coordinate, cellObj.value)
        
#Get the humidity data and append it to Humidity[]
for rowOfCellObjects in sheet['M1' : 'M' + str(sheet.max_row)]:
    for cellObj in rowOfCellObjects:
        if(cellObj.value != None):
        #if isinstance(cellObj.value, long):
            Humidity.append(cellObj.value)
        #print(cellObj.coordinate, cellObj.value)
        
print("** Checking input values with thier corresponding day input ...  ")
print("** Day count: " + str(len(Date)))
print("** Temperature count: " + str(len(Temperature)))
print("** Radon Count: " + str(len(Rn_hourly)))
print("** Air pressure count: " + str(len(Air_Pressure)))
print("** Humidity count: " + str(len(Humidity)))

avg_final = []
dummy_avg = []
dummy_day = []
dummy_min =[]
dummy_max = []
dummy_err = []

def analyzer(obj):
    count = -1 
    ind = Date[0].day
    dummy_day.append(Date[0].strftime("%Y-%m-%d"))
    for row in Date:
        count = count + 1
        dayye = row.day  
        if(dayye == ind):
            if row.strftime("%Y-%m-%d") not in dummy_day:
                dummy_day.append(row.strftime("%Y-%m-%d"))
            count_two = count - 2
            print(str(row) + " : " + str(obj[count_two]))
            count_two = count_two + 1
            dummy_avg.append(obj[count_two])
        else:
            dummy_avg.append(obj[count_two+1])
            print("avg: " + str(np.mean(dummy_avg)))
            avg_final.append(np.mean(dummy_avg)) # calculate the average
            dummy_max.append(np.max(dummy_avg)) # get the max value
            dummy_min.append(np.min(dummy_avg)) # get the min value
            error = (np.std(dummy_avg))/(np.sqrt(len(dummy_avg) - 1)) # get the error
            dummy_err.append(error)
            dummy_day.append(row.strftime("%Y-%m-%d"))
            del dummy_avg[:] #flush out existing data
            ind = dayye
            print("---------------------------- End day data -----------------------")
    return
    
# call the analyzer function
print("** Analyzing data ...")
analyzer(Air_Pressure)

print("** Writing analyzed data to 'data.csv' ...")
with open('data.csv', 'w') as f:
    writer = csv.writer(f)
    writer.writerows(zip(dummy_day,avg_final,dummy_min,dummy_max,dummy_err))
